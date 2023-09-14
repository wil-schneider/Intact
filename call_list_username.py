import openpyxl
import csv
import os
import re
import shutil
from datetime import datetime
import pandas as pd
import argparse

# Define unprocessed_tabs globally
unprocessed_tabs = set()
processed_journal_names = set()
processed_journal_periods = set()

# Load processed files and tabs from the CSV file if it exists
processed_info = {}
processed_file_names = set()

def load_processed_info():
    if os.path.exists('processed_files.csv'):
        with open('processed_files.csv', 'r') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row in csv_reader:
                file_name = row[0]
                processed_tabs = row[1:]
                processed_info[file_name] = processed_tabs
                processed_file_names.add(file_name)

load_processed_info()

# Load processed files and tabs from the CSV file if it exists
processed_info = {}
if os.path.exists('processed_files.csv'):
    with open('processed_files.csv', 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        for row in csv_reader:
            processed_info[row[0]] = set(row[1:])

# Load valid usernames from an Excel file 'valid_usernames.xlsx'
valid_usernames = set()  # Define the set here

# Load valid version number from an Excel file 'valid_version.xlsx'
valid_version = set()  # Define the set here

# Check if the file 'valid_usernames.xlsx' exists
if os.path.exists('valid_usernames.xlsx'):
    workbook = openpyxl.load_workbook('valid_usernames.xlsx', data_only=True, read_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        for cell_value in row:
            valid_usernames.add(cell_value.strip())

# Check if the file 'valid_version.xlsx' exists
if os.path.exists('valid_version.xlsx'):
    workbook = openpyxl.load_workbook('valid_version.xlsx', data_only=True, read_only=True)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        for cell_value in row:
            if cell_value is not None:
                valid_version.add(str(cell_value))  # Convert to string and add to valid_version set

def extract_data_to_csv(source_sheet, start_row, start_col, end_col, dest_csv):
    with open(dest_csv, 'w', newline='') as f:
        c = csv.writer(f)

        # Write the extracted data rows
        for row in source_sheet.iter_rows(min_row=start_row, min_col=start_col, max_col=end_col, values_only=True):
            if any(cell is not None for cell in row):  # Check if any cell in the row has data
                processed_row = [str(cell) if cell is not None else None for cell in row]
                c.writerow(processed_row)

def process_erp_tab(workbook, file_path, sheet_name, output_directory):
    global unprocessed_tabs, processed_journal_names, processed_journal_periods, valid_version, valid_usernames, processed_info

    sheet = workbook[sheet_name]

    journal_name = sheet['C9'].value
    journal_period = datetime.strptime(sheet['C12'].value, '%Y/%m/%d').date()
    version = str(sheet['C6'].value).rstrip('.0')  # Convert to string and remove trailing '.0'
    username = sheet['C8'].value.strip()

    # Extract workbook's name from file_path without the extension
    workbook_name = os.path.splitext(os.path.basename(file_path))[0]

    # Define the output path
    output_file_path = f"C:/Users/sdecarlo/Documents/Projects/Intact/TestingFolder/{workbook_name}_username.txt"
    # Write username to the txt file
    with open(output_file_path, 'w') as file:
        file.write(username)
        
    error_messages = []

    # Check Journal Name
    if journal_name in processed_journal_names:
        error_message = f"Duplicate Journal Name '{journal_name}' in tab '{sheet_name}'."
        error_messages.append(error_message)
    else:
        processed_journal_names.add(journal_name)  # Add journal_name to the set

    # Check Journal Period
    if journal_period in processed_journal_periods:
        error_message = f"Duplicate Journal Period '{journal_period}' in tab '{sheet_name}'."
        error_messages.append(error_message)
    else:
        processed_journal_periods.add(journal_period)  # Add journal_period to the set

    # Check version number
    if version not in valid_version:
        error_message = f"Version '{version}' in tab '{sheet_name}' of file '{file_path}' is not authorized."
        error_messages.append(error_message)

    # Check Username
    if username not in valid_usernames:
        error_message = f"Username '{username}' in tab '{sheet_name}' of file '{file_path}' is not authorized."
        error_messages.append(error_message)

    if error_messages:
        # Print error messages to the terminal
        for error_message in error_messages:
            print(error_message)
        return  # Reject the entire XLSM file

    # Update the processed_info dictionary
    if os.path.basename(file_path) not in processed_info:
        processed_info[os.path.basename(file_path)] = set()
    processed_info[os.path.basename(file_path)].add(journal_name)
    processed_info[os.path.basename(file_path)].add(journal_period)
    
    # Mark Journal Name, Journal Period, and Account Date as processed
    unprocessed_tabs.add(journal_name)
    unprocessed_tabs.add(journal_period)

    # Construct CSV paths using the defined variables
    details_csv_path = os.path.join(output_directory, generate_unique_csv_filename(file_path, f"{sheet_name}_combined", "csv"))
    header_csv_path = os.path.join(output_directory, generate_unique_csv_filename(file_path, f"{sheet_name}_hdr", "csv"))

    # Extract Header Table data (columns B and C, rows 6-14)
    header_data_B = []
    header_data_C = []
    for row in sheet.iter_rows(min_row=6, max_row=14, min_col=2, max_col=3, values_only=True):
        header_data_B.append(re.sub(r'[*\[\]()]', '', str(row[0])) if row[0] is not None else None)
        header_data_C.append(re.sub(r'[*\[\]()]', '', str(row[1])) if row[1] is not None else None)

    # Find the last non-empty column index for each row in the Details table
    start_row = 18
    end_col_indices = []
    for row in sheet.iter_rows(min_row=start_row, min_col=2, max_col=sheet.max_column, values_only=True):
        last_non_empty_col = None
        for col_index, cell in enumerate(row, start=2):  # Start from column B
            if cell is not None:
                last_non_empty_col = col_index
        if last_non_empty_col is not None:
            end_col_indices.append(last_non_empty_col)

    details_data = []
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=2, max_col=max(end_col_indices), values_only=True):
        # Remove special characters [] / \ () * from each cell using a regular expression
        cleaned_row = [re.sub(r'[*\\/().\[\]]', '', str(cell)) if cell is not None else None for cell in row]
        details_data.append(cleaned_row)

    # Combine Header and Details data
    combined_data = []

    # Add the "ProcessedFileName", "TabName", Header data from Column B, and the first row of the Details data
    combined_data.append(["ProcessedFileName", "TabName"] + list(header_data_B) + list(details_data[0]))

    # Add the xlsm file name, tab name, Header data from Column C, and the subsequent rows of the Details data
    for row in details_data[1:]:
        combined_data.append([os.path.basename(file_path), sheet_name] + list(header_data_C) + list(row))

    # Write the combined data to the CSV file
    details_csv_path = os.path.join(output_directory, generate_unique_csv_filename(file_path, f"{sheet_name}_combined", "csv"))
    with open(details_csv_path, 'w', newline='') as f:
        c = csv.writer(f)
        c.writerows(combined_data)

    # Add the processed file to the list of processed files
    processed_file_names.add(os.path.basename(file_path))

    return True  # Indicate that the tab was successfully processed

def generate_unique_csv_filename(xlsm_filename, sheet_name, extension):
    base_filename = f"{os.path.splitext(os.path.basename(xlsm_filename))[0]}_{sheet_name}"
    if xlsm_filename in processed_info:
        processed_files = processed_info[xlsm_filename]
    else:
        processed_files = set()
        processed_info[xlsm_filename] = processed_files
    
    unique_name = f"{base_filename}.{extension}"
    count = 1
    while unique_name in processed_files:
        unique_name = f"{base_filename}_{count}.{extension}"
        count += 1
    processed_files.add(unique_name)
    return unique_name

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]

def combine_csv_files(input_folder, output_folder, output_filename):
    combined_data = []

    # List the CSV filenames in the input_folder and sort them using the natural sorting function
    csv_filenames = [filename for filename in os.listdir(input_folder) if filename.endswith(".csv")]
    csv_filenames.sort(key=natural_sort_key)

    for filename in csv_filenames:
        csv_path = os.path.join(input_folder, filename)

        try:
            # Read the CSV with all columns as strings
            df = pd.read_csv(csv_path, encoding='utf-8-sig', dtype=str)
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(csv_path, encoding='ISO-8859-1', dtype=str)
            except Exception as e:
                print(f"Could not read {filename} due to {e}")
                continue  # Skip this file and move on to the next one

        combined_data.append(df)

    combined_df = pd.concat(combined_data, ignore_index=True)
    output_csv_path = os.path.join(output_folder, output_filename)
    combined_df.to_csv(output_csv_path, index=False)

def validate_workbook(workbook):
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("ERP_"):
            sheet = workbook[sheet_name]
            sheet = workbook[sheet_name]
            journal_name = sheet['C9'].value
            journal_period = sheet['C12'].value
            version = str(sheet['C6'].value).strip()
            username = sheet['C8'].value.strip()

            error_messages = []

            if journal_name in processed_journal_names:
                error_messages.append(f"Duplicate Journal Name '{journal_name}' in tab '{sheet_name}'.")
            if journal_period in processed_journal_periods:
                error_messages.append(f"Duplicate Journal Period '{journal_period}' in tab '{sheet_name}'.")
            if version not in valid_version:
                error_messages.append(f"Version '{version}' in tab '{sheet_name}' is not authorized.")
            if username not in valid_usernames:
                error_messages.append(f"Username '{username}' in tab '{sheet_name}' is not authorized.")

            if error_messages:
                print(f"Validation failed for tab {sheet_name}:")
                for error_message in error_messages:
                    print(f"  - {error_message}")
                return False

    return True

def main(base_directory, file_name, output_folder_name):
    global unprocessed_tabs  # Allow modification of the global set

    # base_directory = input("Enter the base directory: ") 
    # file_name = input("Enter the file name (including .xlsm extension): ")
    # output_folder_name = input("Enter the output folder name: ")

    print(f"Base directory: {base_directory}")
    print(f"File name: {file_name}")
    print(f"Output folder name: {output_folder_name}")

    # Check if the file has been processed before
    if os.path.basename(file_name) in processed_info:
        print(f"Error: File '{file_name}' has already been processed.")
        return

    file_path = os.path.join(base_directory, file_name)
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    print("Validating workbook...")
    if not validate_workbook(workbook):  # Validate the entire workbook first
        print("Validation failed. Aborting processing.")
        return

    # Only proceed with creating subfolders and files if validation succeeds
    output_directory = os.path.join(base_directory, output_folder_name, os.path.splitext(file_name)[0])
    os.makedirs(output_directory, exist_ok=True)

    # Mark the file as processed by adding its name to the dictionary
    processed_info[os.path.basename(file_path)] = set()

    # Process the workbook's tabs and create the necessary CSV files
    tabs_processed = False  # Flag to track whether any tabs were successfully processed
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("ERP_"):
            if process_erp_tab(workbook, file_path, sheet_name, output_directory):
                tabs_processed = True

    if tabs_processed:
        combine_csv_files(output_directory, output_folder_name, os.path.splitext(file_name)[0] + "_combined_data.csv")
    else:
        print("No tabs were successfully processed. Combined CSV will not be generated.")

    # Update the set of processed file names
    processed_file_names.add(os.path.basename(file_path))

    if tabs_processed:
        try:
            shutil.rmtree(output_directory)
            print(f"Deleted subfolder: {output_directory}")
        except Exception as e:
            print(f"Error deleting subfolder: {output_directory}")
            print(f"Error details: {e}")

    # Print the processed file names to the terminal
    for processed_file in processed_file_names:
        print(f"Processed file: {processed_file}")

    # Write the updated list of processed file names to 'processed_files.csv'
    if tabs_processed:
        with open('processed_files.csv', 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            for processed_file in processed_file_names:
                csv_writer.writerow([processed_file])
            
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Your script description")

    # Define the command-line arguments you need
    parser.add_argument("-b", "--base-directory", type=str, required=True, help="The base directory")
    parser.add_argument("-f", "--file-name", type=str, required=True, help="The file name (including .xlsm extension)")
    parser.add_argument("-o", "--output-folder-name", type=str, required=True, help="The output folder name")

    # Parse the arguments
    args = parser.parse_args()

    # Call the main function with the parsed arguments
    main(args.base_directory, args.file_name, args.output_folder_name)